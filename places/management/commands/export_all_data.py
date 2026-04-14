"""
places/management/commands/export_all_data.py
"""

import datetime
from django.core.management.base import BaseCommand
from django.apps import apps
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
ALT_ROW_FILL = PatternFill("solid", start_color="D9E1F2")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Arial", size=9)
CENTER       = Alignment(horizontal="center", vertical="center")
WRAP         = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN         = Border(
    left=Side(style="thin", color="BFBFBF"),   right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),    bottom=Side(style="thin", color="BFBFBF"),
)


def _cell_value(val):
    if val is None:
        return ""
    if isinstance(val, (list, dict)):
        return str(val)
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return str(val)


def _get_fields(model):
    pairs = []
    for f in model._meta.get_fields():
        if f.many_to_many or f.one_to_many:
            continue
        pairs.append((f.name, getattr(f, "attname", f.name)))
    return pairs


def _write_sheet(wb, sheet_name, model):
    ws     = wb.create_sheet(title=sheet_name[:31])
    fields = _get_fields(model)
    qs     = model.objects.all()
    count  = qs.count()

    for ci, (name, _) in enumerate(fields, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font, c.fill, c.alignment, c.border = HEADER_FONT, HEADER_FILL, CENTER, THIN

    for ri, obj in enumerate(qs.iterator(), 2):
        fill = ALT_ROW_FILL if ri % 2 == 0 else None
        for ci, (name, attname) in enumerate(fields, 1):
            raw = getattr(obj, attname, getattr(obj, name, ""))
            c   = ws.cell(row=ri, column=ci, value=_cell_value(raw))
            c.font, c.alignment, c.border = BODY_FONT, WRAP, THIN
            if fill:
                c.fill = fill

    for ci, (name, _) in enumerate(fields, 1):
        col   = get_column_letter(ci)
        width = len(name)
        for ri in range(2, min(count + 2, 202)):
            v = ws.cell(row=ri, column=ci).value
            if v:
                width = max(width, min(len(str(v)), 60))
        ws.column_dimensions[col].width = width + 2

    ws.freeze_panes = "A2"
    note = ws.cell(row=1, column=len(fields) + 2, value=f"{count} rows")
    note.font      = Font(name="Arial", italic=True, size=8, color="888888")
    note.alignment = CENTER
    return count


def _write_cover(wb, summary):
    ws = wb.create_sheet(title="Index", index=0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    ws.cell(row=1, column=1, value="Database Backup").font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws.cell(row=2, column=1, value=f"Exported: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(name="Arial", size=9, color="888888")

    for col, hdr in [(1, "Sheet / Model"), (2, "Row Count")]:
        ws.cell(row=4, column=col, value=hdr).font = Font(name="Arial", bold=True, size=10)

    for i, (name, count) in enumerate(summary, 5):
        ws.cell(row=i, column=1, value=name).font  = Font(name="Arial", size=9)
        ws.cell(row=i, column=2, value=count).font = Font(name="Arial", size=9)
        if i % 2 == 0:
            for col in (1, 2):
                ws.cell(row=i, column=col).fill = ALT_ROW_FILL

    tr = len(summary) + 5
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=9)
    ws.cell(row=tr, column=2, value=f"=SUM(B5:B{tr-1})").font = Font(name="Arial", bold=True, size=9)
    ws.freeze_panes = "A5"


class Command(BaseCommand):
    help = "Export all database models to a single .xlsx backup file"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output", "-o",
            default=f"db_backup_{datetime.date.today()}.xlsx",
            help="Output filename (default: db_backup_YYYY-MM-DD.xlsx)",
        )

    def handle(self, *args, **options):
        output = options["output"]

        app_label = None
        for app_config in apps.get_app_configs():
            try:
                app_config.get_model("Place")
                app_label = app_config.label
                break
            except LookupError:
                pass

        if not app_label:
            self.stderr.write(self.style.ERROR("Could not detect app label — is 'Place' registered?"))
            return

        MODEL_REGISTRY = [("Users", User)]
        for name in [
            "ExpertArea", "Category",
            "Place", "PlaceImage", "PlaceVideo",
            "CheckIn",
            "Trail", "TrailPlace", "TrailCompletion",
            "Comment", "Vote",
            "Favorite", "TrailFavorite",
            "Badge", "UserBadge",
            "Challenge", "UserChallengeCompletion",
            "Notification",
            "TourOffering", "TourPackage", "TourItineraryDay",
            "UserProfile",
        ]:
            try:
                MODEL_REGISTRY.append((name, apps.get_model(app_label, name)))
            except LookupError:
                self.stdout.write(self.style.WARNING(f"  ⚠  '{name}' not found — skipping"))

        wb      = Workbook()
        wb.remove(wb.active)
        summary = []

        self.stdout.write(f"\nExporting {len(MODEL_REGISTRY)} models → {output}\n")

        for sheet_name, model in MODEL_REGISTRY:
            try:
                count = _write_sheet(wb, sheet_name, model)
                summary.append((sheet_name, count))
                self.stdout.write(f"  {self.style.SUCCESS('✓')}  {sheet_name:<28} {count:>6,} rows")
            except Exception as exc:
                self.stderr.write(f"  ✗  {sheet_name:<28} ERROR: {exc}")
                summary.append((sheet_name, "ERROR"))

        _write_cover(wb, summary)
        wb.save(output)

        total = sum(c for _, c in summary if isinstance(c, int))
        self.stdout.write(f"\n{self.style.SUCCESS('Saved:')} {output}")
        self.stdout.write(f"Sheets: {len(summary)}  |  Total rows: {total:,}\n")